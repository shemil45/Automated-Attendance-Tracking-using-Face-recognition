"""
Attendance Report Generator
Generates Excel reports from attendance CSV logs
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ReportGenerator:
    def __init__(self):
        """Initialize report generator"""
        self.base_dir = Path(__file__).parent.parent
        self.attendance_dir = self.base_dir / "attendance"
        self.students_file = self.base_dir / "data" / "students.csv"
        
    def get_attendance_files(self):
        """Get list of attendance CSV files"""
        if not self.attendance_dir.exists():
            return []
        
        files = list(self.attendance_dir.glob("attendance_*.csv"))
        return sorted(files, reverse=True)  # Most recent first
    
    def load_students(self):
        """Load student database"""
        if not self.students_file.exists():
            print(f"Error: Student database not found: {self.students_file}")
            return None
        
        df = pd.read_csv(self.students_file)
        return df
    
    def load_attendance(self, attendance_file):
        """Load attendance log"""
        df = pd.read_csv(attendance_file)
        return df
    
    def generate_report(self, attendance_file):
        """Generate Excel report from attendance CSV"""
        print(f"\nGenerating report from: {attendance_file.name}")
        
        # Load data
        students_df = self.load_students()
        if students_df is None:
            return None
        
        attendance_df = self.load_attendance(attendance_file)
        
        # Merge data
        # Mark all students as Absent first
        report_df = students_df.copy()
        report_df['Status'] = 'Absent'
        report_df['Timestamp'] = '-'
        
        # Update with present students
        for _, row in attendance_df.iterrows():
            regno = row['RegNo']
            mask = report_df['RegNo'] == regno
            report_df.loc[mask, 'Status'] = 'Present'
            report_df.loc[mask, 'Timestamp'] = row['Timestamp']
        
        # Sort by RegNo
        report_df = report_df.sort_values('RegNo')
        
        # Calculate statistics
        total_students = len(report_df)
        present_count = len(report_df[report_df['Status'] == 'Present'])
        absent_count = total_students - present_count
        attendance_rate = (present_count / total_students * 100) if total_students > 0 else 0
        
        # Extract session info from filename
        filename = attendance_file.stem  # attendance_20260105_143000
        date_str = filename.split('_')[1]  # 20260105
        time_str = filename.split('_')[2]  # 143000
        
        session_date = datetime.strptime(date_str, '%Y%m%d').strftime('%Y-%m-%d')
        session_time = datetime.strptime(time_str, '%H%M%S').strftime('%H:%M:%S')
        
        # Create Excel file
        excel_file = attendance_file.with_suffix('.xlsx')
        
        # Write to Excel with pandas
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Write main data
            report_df.to_excel(writer, sheet_name='Attendance', index=False, startrow=8)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Attendance']
            
            # Add header information
            worksheet['A1'] = 'ATTENDANCE REPORT'
            worksheet['A1'].font = Font(size=16, bold=True)
            
            worksheet['A3'] = 'Session Date:'
            worksheet['B3'] = session_date
            worksheet['A4'] = 'Session Time:'
            worksheet['B4'] = session_time
            
            worksheet['A6'] = 'Total Students:'
            worksheet['B6'] = total_students
            worksheet['A7'] = 'Present:'
            worksheet['B7'] = f"{present_count} ({attendance_rate:.1f}%)"
            worksheet['C7'] = 'Absent:'
            worksheet['D7'] = f"{absent_count} ({100-attendance_rate:.1f}%)"
            
            # Format header cells
            for cell in ['A1', 'A3', 'A4', 'A6', 'A7', 'C7']:
                worksheet[cell].font = Font(bold=True)
            
            # Format data table
            # Header row (row 9)
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for col in range(1, len(report_df.columns) + 1):
                cell = worksheet.cell(row=9, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data rows
            for row_idx in range(10, 10 + len(report_df)):
                for col_idx in range(1, len(report_df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Color code status
                    if col_idx == report_df.columns.get_loc('Status') + 1:
                        if cell.value == 'Present':
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                            cell.font = Font(color='006100', bold=True)
                        else:  # Absent
                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                            cell.font = Font(color='9C0006', bold=True)
            
            # Adjust column widths
            column_widths = {
                'A': 15,  # RegNo
                'B': 20,  # Name
                'C': 15,  # Class
                'D': 30,  # Email
                'E': 12,  # Status
                'F': 20,  # Timestamp
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
        
        print(f"✓ Report generated: {excel_file}")
        print(f"\nSummary:")
        print(f"  Total Students: {total_students}")
        print(f"  Present: {present_count} ({attendance_rate:.1f}%)")
        print(f"  Absent: {absent_count} ({100-attendance_rate:.1f}%)")
        
        return excel_file
    
    def run(self):
        """Main report generation function"""
        print("\n" + "="*60)
        print("Attendance Report Generator")
        print("="*60)
        
        # Get attendance files
        files = self.get_attendance_files()
        
        if not files:
            print("\n⚠ No attendance files found in attendance/ directory")
            print("Please run attendance_system.py first to create attendance logs")
            return
        
        print(f"\nFound {len(files)} attendance file(s):")
        for i, file in enumerate(files, 1):
            print(f"  {i}. {file.name}")
        
        # Select file
        if len(files) == 1:
            selected_file = files[0]
            print(f"\nUsing: {selected_file.name}")
        else:
            while True:
                try:
                    choice = input(f"\nSelect file number (1-{len(files)}): ").strip()
                    idx = int(choice) - 1
                    
                    if 0 <= idx < len(files):
                        selected_file = files[idx]
                        break
                    else:
                        print(f"Invalid choice. Please enter 1-{len(files)}")
                except ValueError:
                    print("Please enter a valid number")
        
        # Generate report
        excel_file = self.generate_report(selected_file)
        
        if excel_file:
            print(f"\n{'='*60}")
            print(f"✓ Report saved to: {excel_file}")
            print(f"{'='*60}\n")

def main():
    """Main function"""
    generator = ReportGenerator()
    generator.run()

if __name__ == "__main__":
    main()
